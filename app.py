from flask import Flask, render_template, request, session # import flask libraries
import openpyxl #a Python library for reading and writing Excel (with extension xlsx/xlsm/xltx/xltm) files
import mysql.connector

# Assume this is the excel file:
# Brand     Model       Color	Year
# Toyota    Corolla     Red	    2005
# Honda     Civic       Blue	2006
# Ford      Mustang     ellow	2007
# Chevrolet Camaro      Black	2008
# BMW       M3          White	2009

# In XAMPP MySQL,   
# table car:
#   brand varchar(64)
#   model varchar(64)
#   color varchar(64)
#   year int(4)

#initialize sql database connector
db = mysql.connector.connect(
  host="localhost",
  user="root",
  password="",
  database="test" #database name
)

app = Flask(__name__)
app.secret_key = 'secret_key' #secret_key for session method

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods = ['POST'])
def upload():
    uploadedFile = request.files['file'] # xlsx/xlsm/xltx/xltm files (not .csv)

    excelWorkbook = openpyxl.load_workbook(uploadedFile) #load the excel workbook using openpyxl
    workbookSheet = excelWorkbook.active #get the active worksheet from workbook
    
    excelColumnNames = []
    excelData = []

    # get max number of columns and rows
    maxColumns = workbookSheet.max_column
    maxRows = workbookSheet.max_row

    for col in range(1, maxColumns + 1): #loop from 1 to max num of columns
                                         #excel starts with 1 and not 0
        column_data = []
        
        for row in range(1, maxRows + 1): #loop from 1 to max num of rows
            cell = workbookSheet.cell(row=row, column=col) #get data of cell at current row and column

            if row == 1: #if first row of the column (the column name/title)
                excelColumnNames.append(cell.value)

            column_data.append(cell.value if cell.value is not None else "") #if cell is not empty, append value. else append blank
        
        excelData.append(column_data) #after all rows of current column, append it to array

    session['excelData'] = excelData #saves excelData into a session so it can be used in another function
    # excelData:  [['Brand', 'Toyota', 'Honda', 'Ford', 'Chevrolet', 'BMW'], 
    # ['Model', 'Corolla', 'Civic', 'Mustang', 'Camaro', 'M3'], 
    # ['Color', 'Red', 'Blue', 'Yellow', 'Black', 'White'], 
    # ['Year', 2005, 2006, 2007, 2008, 2009]]
    
    cursor = db.cursor()
    cursor.execute("SELECT * FROM car") #from table: car
    cursor.fetchall()
    
    SQLColumnNames = []
    for i in cursor.description: #gets the column name in cursor.description
        SQLColumnNames.append(i[0])
    # SQLColumnNames:  ['brand', 'model', 'color', 'year']

    SQLColumnNamesLength = len(SQLColumnNames)
    # SQLColumnNamesLength: 4

    return render_template('upload.html', excelColumnNames=excelColumnNames, SQLColumnNames=SQLColumnNames, SQLColumnNamesLength=SQLColumnNamesLength)

@app.route('/insert', methods = ['POST'])
def insert():
    excelData = session.get('excelData', []) #gets the excelData from the other function

    excelDict = {}

    for excelColumnName in request.form:
        # request.form:  ImmutableMultiDict([('Brand', 'brand'), ('Model', 'model'), 
        # ('Color', 'color'), ('Year', 'year')])
        
        print("================LOOP 1===================")
        selectedSQLColumnName = request.form[excelColumnName]
        # excelColumnName: Brand
        # selectedSQLColumnName: brand
        
        if selectedSQLColumnName != "None": #if a column is selected
            for excelColumn in excelData:
                print("================LOOP 2===================")
                print("excelColumn: ", excelColumn)
                # excelColumn:  ['Brand', 'Toyota', 'Honda', 'Ford', 'Chevrolet', 'BMW']
                if excelColumn[0] == excelColumnName:
                    print("excelColumn[1:]: ", excelColumn[1:])
                    # excelColumn[1:]:  ['Toyota', 'Honda', 'Ford', 'Chevrolet', 'BMW']

                    excelDict[selectedSQLColumnName] = excelColumn[1:]
                    # excelDict:  {'brand': ['Toyota', 'Honda', 'Ford', 'Chevrolet', 'BMW']}
        
    dataToBeInsertedDict = {}
    for key in excelDict:
        print(key)
        
    for i in range(len(excelDict[key])): #len(excelDict[key]) --> # of columns to be inserted in SQL
        # gets first values of every column, then gets the second values of all columns, and so on...
        for SQLColumnName in excelDict:
            dataToBeInsertedDict[SQLColumnName] = excelDict[SQLColumnName][i]
            # SQLColumnName: brand
            # excelDict[columnName]:  ['Toyota', 'Honda', 'Ford', 'Chevrolet', 'BMW']
            # excelDict[columnName][0]:  Toyota

        print("dataToBeInsertedDict: ", dataToBeInsertedDict)
        # dataToBeInsertedDict:  {'brand': 'Toyota', 'model': 'Corolla', 'color': 'Red', 'year': 2005}
        # the first values of all columns to be inserted to the first row of the SQL table

        columns = ""
        values = ""
        data = []

        for key, value in dataToBeInsertedDict.items():
            
            if columns != "": #if not empty
                columns += ", " #add comma and space
                values += ", " #add comma and space

            columns += key # Add the key to the columns string (key is SQLColumnName)

            values += "%s" # string placeholder needed for SQL statement syntax

            data.append(value) # Add the value to the data list

        sql = "INSERT INTO car (" + columns + ") VALUES (" + values + ")"
        # sql = "INSERT INTO table (column1, column2) VALUES (%s, %s)" 
        # data = ("value1", "value2") == VALUES (%s, %s)

        data = tuple(data) # Convert the data list to a tuple

        cursor = db.cursor()
        cursor.execute(sql, data)
        db.commit()
                        
    return render_template('insert.html')

app.run(debug=True)